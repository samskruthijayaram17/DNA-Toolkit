import os
import csv
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
fasta_file = os.path.join(script_dir, "data", "example_sequences.fasta")
sequences = []
if os.path.exists(fasta_file):
    with open(fasta_file) as f:
        sequence_name = None
        sequence = ""
        for line in f:
            line = line.strip()  
            if not line:  
                continue
            if line.startswith(">"): 
                if sequence_name is not None:
                    sequences.append((sequence_name, sequence))
                sequence_name = line[1:]
                sequence = ""
            else:  
                sequence += line
        if sequence_name is not None:
            sequences.append((sequence_name, sequence))
else:
    print(f"Warning: FASTA file not found at {fasta_file}")

print("\n--- SEQUENCE ANALYSIS ---")
def gc_content(DNA):
    gc=0
    for base in DNA:
        if base == 'G' or base == 'C':
            gc+=1
    percentage=(gc/len(DNA))*100
    return percentage

def detect_mutations(ref, sample):
    mutations = []
    length = min(len(ref), len(sample))
    for i in range(length):
        if ref[i] != sample[i]:
            mutations.append((i+1, ref[i], sample[i]))
    return mutations
print("Sequences loaded:")
for name, seq in sequences:
    print(f"  {name}: {seq}")
if len(sequences) < 2:
    print("\nError: Need at least 2 sequences (1 reference + 1 sample)")
else:
    # First sequence is reference
    reference_name, reference_seq = sequences[0]
    print(f"\n--- REFERENCE SEQUENCE (seq1: {reference_name}) ---")
    print(f"{reference_seq}")
    
    # All other sequences are samples
    print(f"\n--- MUTATION DETECTION RESULTS ---")
    for i in range(1, len(sequences)):
        sample_name, sample_seq = sequences[i]
        mutations = detect_mutations(reference_seq, sample_seq)
        
        print(f"\nSample {i} ({sample_name}):")
        print(f"Sequence: {sample_seq}")
        
        if mutations:
            print(f"Mutations found: {len(mutations)}")
            for pos, ref_base, sample_base in mutations:
                print(f"  Position {pos}: {ref_base} -> {sample_base}")
        else:
            print("No mutations detected")

def find_motif(sequence, query):
    positions = []
    for i in range(len(sequence) - len(query) + 1):
        if sequence[i:i+len(query)] == query:
            positions.append(i+1)
    return positions
print("\n--- MOTIF SEARCH ---")
query = "GC"  # set motif herequery 
query = (query or "").strip().upper()
    
if not query:
        print("No motif provided in code. Skipping motif search.")
else:
        print(f"\nSearching for motif: {query}\n")
        for i in range(1, len(sequences)):
            sample_name, sample_seq = sequences[i]
            sample_seq = sample_seq.upper()
            matches = find_motif(sample_seq, query)
            print(f"\nSample {i} ({sample_name}):")
            if matches:
                print(f"  Positions found: {matches}")
            else:
                print(f"  No matches found")

print("\n--- STATISTICAL ANALYSIS ---")
def find_csv_file(filename="gene_expression.csv"):
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Search in examples folder and sibling data folder
    search_paths = [
        os.path.join(script_dir, "examples", filename),
        os.path.join(script_dir, "..", "data", filename),
        os.path.join(script_dir, "data", filename)
    ]
    
    for path in search_paths:
        if os.path.exists(path):
            return path
            
    if filename.endswith(".csv"):
        xlsx_filename = filename + ".xlsx"
        for path in search_paths:
            xlsx_path = path + ".xlsx"
            if os.path.exists(xlsx_path):
                return xlsx_path
    raise FileNotFoundError(f"{filename} (or .xlsx variant) not found in examples folder")
def read_data_file(filepath):
    genes_data = []
    if filepath.endswith('.xlsx'):
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            gene_idx = headers.index("gene") if "gene" in headers else 0
            expr_idx = headers.index("expression") if "expression" in headers else 1
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[gene_idx] is not None:
                    genes_data.append((row[gene_idx], row[expr_idx]))    
        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}")
    else:
        try:
            with open(filepath) as f:
                reader = csv.DictReader(f)
                fieldnames = reader.fieldnames or []
                gene_col = "gene_id" if "gene_id" in fieldnames else "gene"
                expr_col = "sample1" if "sample1" in fieldnames else "expression"
                for row in reader:
                    if row.get(gene_col):
                        genes_data.append((row[gene_col], row[expr_col]))
        except Exception as e:
            raise Exception(f"Error reading CSV file: {e}")
    return genes_data

csv_file = None
try:
    csv_file = find_csv_file()
    genes_data = []
    total = 0
    count = 0
    max_gene = ""
    max_value = 0
    raw_data = read_data_file(csv_file)
    for gene, expression in raw_data:
        try:
            expr_value = float(expression)
            genes_data.append((gene, expr_value))
            total += expr_value
            count += 1
            if expr_value > max_value:
                max_value = expr_value
                max_gene = gene    
        except ValueError:
            print(f"Warning: Could not convert expression to float for {gene}")
    if count == 0:
        print("Error: No valid gene expression data found")
    else:
        print("Gene expression data:")
        for gene, expression in genes_data:
            print(f"  {gene}: {expression}")
        average = total / count
        print(f"\nAverage expression: {average:.5f}")
        print(f"Highest expressed gene: {max_gene}- {max_value:.5f}")
except FileNotFoundError as e:
    print(f"Error: {e}")
except Exception as e:
    print(f"Unexpected error: {e}")

if csv_file:
    df = pd.read_csv(csv_file)
    high_expression = df[df["sample1"] > 10] 
    print("\nHigh expression genes:")
    print(high_expression)
    print("\nImportant gene:")
    for gene in high_expression["gene_id"]:
        print(gene)

    plt.figure(figsize=(10, 5))
    plt.bar(df["gene_id"], df["sample1"])
    plt.xlabel("Gene ID")
    plt.ylabel("Expression Level")
    plt.title("Gene Expression Levels")
    plt.xticks(rotation=45)
    plt.show()

    plt.figure(figsize=(8, 6))
    plt.imshow(df[["sample1", "sample2"]], aspect="auto", cmap="Blues")
    plt.title("Gene Expression Heatmap")
    plt.yticks(range(len(df["gene_id"])), df["gene_id"])
    plt.colorbar()
    plt.show()